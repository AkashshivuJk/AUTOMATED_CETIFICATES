from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import os

# Step 1: Load the Excel file
excel_file = 'NAMES_EXCEL_FILE.xlsx'  # Replace with your Excel file name
wb = load_workbook(excel_file)
sheet = wb.active

# Step 2: Load the certificate template
template_file = 'YOUR_CERTIFICATE_TEMPLATE.png'  # Replace with your template file name
template = Image.open(template_file)

# Step 3: Configure font and text placement
font_file = 'Times New Roman Bold.ttf'  # Replace with the correct path to your font file
font_size = 50  # Adjust font size based on your certificate template
font = ImageFont.truetype(font_file, font_size)  # Load the font

# Define the position where the name should appear
name_position = (500, 490)  # Adjust based on your template

# Step 4: Create an output folder for certificates
output_folder = 'certificates'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Step 5: Generate certificates
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    participant_name = row[0].value  # Get the participant's name
    
    # Copy the template for each participant
    certificate = template.copy()
    draw = ImageDraw.Draw(certificate)
    
    # Add the participant's name to the template
    draw.text(name_position, participant_name, font=font, fill="black")  # Adjust color if needed
    
    # Save the certificate
    certificate_file = os.path.join(output_folder, f'certificate_{participant_name}.png')
    certificate.save(certificate_file)

    print(f'Certificate generated for {participant_name}.')

print("All certificates generated successfully!")
